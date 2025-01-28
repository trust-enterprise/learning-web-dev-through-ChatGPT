import openpyxl

def rearrange_excel_data(file_path):
    # Load the workbook and sheets
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb["Sheet1"]

    # Create SheetB if it doesn't exist
    if "SheetB" not in wb.sheetnames:
        sheetb = wb.create_sheet("SheetB")
    else:
        sheetb = wb["SheetB"]

    # Clear any existing data in SheetB
    for row in sheetb["A1:Z1000"]:
        for cell in row:
            cell.value = None

    # Write headers to SheetB
    sheetb["A1"] = "Title"
    sheetb["B1"] = "Time"
    sheetb["C1"] = "Details"

    # Process data from Sheet1
    dest_row = 2
    for i in range(1, 31, 3):  # Rows 1, 4, 7,... contain titles
        title = sheet1[f"A{i}"].value
        time = sheet1[f"A{i+1}"].value
        details = sheet1[f"A{i+2}"].value

        # Write to SheetB
        sheetb[f"A{dest_row}"] = title
        sheetb[f"B{dest_row}"] = time
        sheetb[f"C{dest_row}"] = details

        dest_row += 1

    # Save the workbook
    wb.save(file_path)
    print("Data rearranged successfully!")

# Provide the file path to the src.xlsx file
file_path = r"C:\Users\lab I\Desktop\test\src.xlsx"
rearrange_excel_data(file_path)
